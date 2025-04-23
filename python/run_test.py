import requests
import json
import csv
import time
import threading
import os
import re
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Backend-Konfiguration
BACKEND_URL = "http://localhost:8000/full-check"

# Verzeichnisse
REPORT_DIR = "C:/Users/bfranneck/Desktop/Praxisprojekt/Anwendung/python/reports"
EXCEL_DIR = "C:/Users/bfranneck/Desktop/Praxisprojekt/Anwendung/python/excel"
os.makedirs(REPORT_DIR, exist_ok=True)
os.makedirs(EXCEL_DIR, exist_ok=True)

def sanitize_url(url: str) -> str:
    clean_url = re.sub(r"https?://", "", url)
    return re.sub(r"\W+", "_", clean_url).strip("_")

def loading_animation(stop_event):
    dots = ""
    while not stop_event.is_set():
        print(f"\rTest wird ausgeführt. Bitte Geduld{dots}  ", end="", flush=True)
        dots += "." if len(dots) < 3 else ""
        if len(dots) > 2:
            dots = ""
        time.sleep(0.5)

def get_priority(impact: str) -> str:
    return {
        "critical": "Hoch",
        "serious": "Hoch",
        "moderate": "Mittel",
        "minor": "Niedrig"
    }.get(impact, "Unbekannt")

# Benutzer-Eingabe
try:
    target_url = input("🌐 Bitte zu scannende URL eingeben (inkl. https://): ").strip()
except KeyboardInterrupt:
    print("\nAbgebrochen.")
    exit(0)

if not target_url.startswith("http"):
    print("❌ Ungültige URL. Bitte mit https:// beginnen.")
    exit(1)

# Ladeanimation starten
stop_event = threading.Event()
loader_thread = threading.Thread(target=loading_animation, args=(stop_event,))
loader_thread.start()

# Anfrage senden
try:
    response = requests.post(BACKEND_URL, json={"url": target_url}, timeout=1000)
    response.raise_for_status()
    result = response.json()
    pages = result.get("results") or result.get("pages") or []
except Exception as e:
    stop_event.set()
    loader_thread.join()
    print(f"\n❌ Fehler bei der Anfrage: {e}")
    exit(1)

# Ladeanimation beenden
stop_event.set()
loader_thread.join()
print("\n✅ Scan abgeschlossen. Ergebnisse werden gespeichert...\n")

# Dateien benennen
timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M")
base_name = f"scan_{sanitize_url(target_url)}_{timestamp}"
json_file = os.path.join(REPORT_DIR, f"{base_name}.json")
csv_file = os.path.join(REPORT_DIR, f"{base_name}.csv")
xlsx_file = os.path.join(EXCEL_DIR, f"{base_name}.xlsx")

if not pages:
    print("⚠️ Keine Ergebnisse erhalten.")
else:
    # JSON speichern (optional)
    with open(json_file, "w", encoding="utf-8") as f:
        json.dump(pages, f, indent=2, ensure_ascii=False)

    # CSV speichern (optional)
    with open(csv_file, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["URL", "AXE Errors", "Structural Issues", "CSS Issues", "ARIA Issues", "Warnings", "Total Errors"])
        for entry in pages:
            summary = entry.get("summary", {})
            writer.writerow([
                entry.get("url", "-"),
                summary.get("axe_errors", 0),
                summary.get("structural_issues", 0),
                summary.get("css_issues", 0),
                summary.get("aria_issues", 0),
                summary.get("warnings", 0),
                summary.get("total_errors", 0),
            ])

    # Excel-Datei mit deduplizierten Einträgen erstellen
    wb = Workbook()
    ws = wb.active
    ws.title = "Accessibility-Report"

    headers = [
        "Seiten-URL", "Fehlertyp", "Priorität", "Beschreibung",
        "HTML-Auszug", "Ziel-Element", "Kategorie", "Hilfelink"
    ]
    ws.append(headers)

    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    seen = set()  # Neue Datenstruktur zum Deduplizieren

    for page in pages:
        page_url = page.get("url", "")
        result = page.get("result", {})

        for issue_type, label in [
            ("axe_violations", "AXE-Fehler"),
            ("css_issues", "CSS-Fehler"),
            ("incomplete_warnings", "Unvollständig")
        ]:
            issues = result.get(issue_type, [])
            if isinstance(issues, list):
                for issue in issues:
                    if issue_type == "css_issues":
                        key = (
                            page_url,
                            label,
                            "Niedrig",
                            issue.get("message", ""),
                            issue.get("snippet", ""),
                            issue.get("selector", ""),
                            "CSS",
                            ""
                        )
                        if key not in seen:
                            ws.append(list(key))
                            seen.add(key)
                    else:
                        for node in issue.get("nodes", []):
                            key = (
                                page_url,
                                label,
                                get_priority(issue.get("impact", "minor")),
                                issue.get("description", ""),
                                node.get("html", ""),
                                ", ".join(node.get("target", [])),
                                ", ".join(issue.get("tags", [])),
                                issue.get("help_url", "")
                            )
                            if key not in seen:
                                ws.append(list(key))
                                seen.add(key)

    for col in ws.columns:
        max_len = max((len(str(cell.value)) if cell.value else 0) for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 5, 60)

    wb.save(xlsx_file)

    print(f"📊 Überprüfung der Barrierefreiheit abgeschlossen.\nEine Zusammenfassung wurde im Ordner 'Excel' gespeichert:\n- {xlsx_file}")

input("🟢 Bitte Taste drücken, um das Programm zu beenden.")
